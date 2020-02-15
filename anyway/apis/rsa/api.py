import json, os

from anyway import db
from anyway.core.utils import Utils
from anyway.core.constants import CONST
from anyway.apis.common.models.cbs_models import AccidentMarker

from flask_restplus import Namespace, Resource, fields
from openpyxl import load_workbook
from dateutil import parser as date_parser

rsa_api = Namespace('rsa', description='RSA related operations')

load_model = rsa_api.model("Model", {"message": fields.String})


@rsa_api.route('/load', endpoint='load')
@rsa_api.response(404, 'Endpoint was not found')
class RSA(Resource):
    parser = rsa_api.parser()
    parser.add_argument('filename', type=str, location='args')

    @rsa_api.param('filename', 'RSA file path')
    @rsa_api.expect(parser)
    @rsa_api.marshal_with(load_model)
    def get(self):
        args = self.parser.parse_args()
        self.parse(args['filename'])
        return {"message": "The file was uploaded successfully"}

    def _iter_rows(self, filename):
        workbook = load_workbook(filename, read_only=True)
        sheet = workbook[u"Worksheet1"]
        rows = sheet.rows
        first_row = next(rows)
        headers = [u'מזהה', u'סוג עבירה', u'סוג רכב', u'נ״צ', u'סרטון', u'תאריך דיווח']
        assert [cell.value for cell in first_row] == headers
        for row in rows:
            id_ = int(row[0].value)
            provider_and_id_ = int(str(CONST.RSA_PROVIDER_CODE) + str(id_))

            q = db.session.query(AccidentMarker).filter(AccidentMarker.id.in_([id_]))
            if q.all():
                q.delete(synchronize_session='fetch')
                db.session.commit()

            violation = row[1].value
            vehicle_type = row[2].value
            coordinates = row[3].value
            video_link = row[4].value
            timestamp = date_parser.parse(row[5].value, dayfirst=True)
            if not violation:
                continue

            vehicle_type = vehicle_type or ''
            coordinates = coordinates.split(',')
            latitude, longitude = float(coordinates[0]), float(coordinates[1])
            description = {'VIOLATION_TYPE': violation, 'VEHICLE_TYPE': vehicle_type}

            yield {'id': id_,
                   'provider_and_id': provider_and_id_,
                   'latitude': latitude,
                   'longitude': longitude,
                   'created': timestamp,
                   'provider_code': CONST.RSA_PROVIDER_CODE,
                   'accident_severity': 0,
                   'title': 'שומרי הדרך',
                   'description': json.dumps(description),
                   'location_accuracy': 1,
                   'type': CONST.MARKER_TYPE_ACCIDENT,
                   'video_link': video_link,
                   'vehicle_type_rsa': vehicle_type,
                   'violation_type_rsa': violation,
                   'accident_year': timestamp.year}

    def parse(self, filename):
        if os.path.exists(filename):
            for batch in Utils.batch_iterator(self._iter_rows(filename), batch_size=50):
                db.session.bulk_insert_mappings(AccidentMarker, batch)
                db.session.commit()

            """
            Fills empty geometry object according to coordinates in database
            """
            db.session.execute('UPDATE markers SET geom = ST_SetSRID(ST_MakePoint(longitude,latitude),4326)\
                                   WHERE geom IS NULL;')
            db.session.commit()
        else:
            rsa_api.abort(404)
